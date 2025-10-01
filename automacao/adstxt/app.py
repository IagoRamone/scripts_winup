import pyautogui as py
import time
import pandas as pd
from openpyxl import load_workbook

py.PAUSE = 1 

py.press("win")
py.write("Google Chrome")
py.press("enter")
time.sleep(1) 

df = pd.read_excel("teste.xlsx")  

arquivo_excel = "teste.xlsx"

wb = load_workbook(arquivo_excel)
ws = wb.active

for index, row in df.iterrows():
    valor = row[0]  
    py.write(str(valor)) 
    py.press("enter")
    time.sleep(2)
    
    py.hotkey("ctrl", "t")
    time.sleep(1)
    py.write(str(valor))
    py.press("enter")
    py.hotkey("ctrl", "w")
    py.hotkey("ctrl", "t")
    py.write(str(valor))
    py.press("enter")
    
    ws[f"B{index + 2}"] = "✅" 

# Salva a planilha
wb.save(arquivo_excel)
