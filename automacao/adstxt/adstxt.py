import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from urllib.parse import urlparse
import time
import os

def processar_arquivo(caminho_excel, progress_var, log_box, btn_iniciar):
    try:
        df = pd.read_excel(caminho_excel, skiprows=1, header=None)
        wb = load_workbook(caminho_excel)
        ws = wb.active

        options = webdriver.ChromeOptions()
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--headless")  # roda em background

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        wait = WebDriverWait(driver, 20)

        total = len(df)
        for index, row in df.iterrows():
            linha_excel = index + 2  
            url = str(row[0]).strip()
            if not url or url.lower() in ("nan", "none"):
                ws[f"B{linha_excel}"] = "(vazio)"
                ws[f"D{linha_excel}"] = "N/A"
                log_box.insert(tk.END, f"[{linha_excel}] Linha vazia\n")
                progress_var.set((index + 1) / total * 100)
                continue

            if not urlparse(url).scheme:
                url = "http://" + url

            if not url.endswith("/ads.txt"):
                if url.endswith("/"):
                    url = url + "ads.txt"
                else:
                    url = url + "/ads.txt"

            try:
                driver.get(url)
                wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                time.sleep(1)

                body_text = driver.find_element(By.TAG_NAME, "body").text.strip()
                lines = [l for l in (ln.strip() for ln in body_text.splitlines()) if l]
                primeiro_texto = lines[0] if lines else ""

                ws[f"B{linha_excel}"] = primeiro_texto
                ws[f"C{linha_excel}"] = (body_text[:32000] if body_text else "")
                ws[f"D{linha_excel}"] = "OK"
                log_box.insert(tk.END, f"[{linha_excel}] OK - {url}\n")

            except Exception as e:
                ws[f"B{linha_excel}"] = f"ERRO: {str(e)[:200]}"
                ws[f"D{linha_excel}"] = "ERROR"
                log_box.insert(tk.END, f"[{linha_excel}] ERRO - {url}\n")

            # Atualiza progresso visual
            progress_var.set((index + 1) / total * 100)
            root.update_idletasks()

        wb.save(caminho_excel)
        driver.quit()

        messagebox.showinfo("Finalizado", "Checklist finalizado com sucesso!")
        log_box.insert(tk.END, "\n✅ Processo concluído! Planilha salva e atualizada.\n")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")
        log_box.insert(tk.END, f"❌ Erro fatal: {e}\n")
    finally:
        btn_iniciar.config(state="normal")

# ------------------- FUNÇÕES DA INTERFACE -------------------
def escolher_arquivo():
    caminho = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Planilhas Excel", "*.xlsx")]
    )
    if caminho:
        caminho_var.set(caminho)
        label_arquivo.config(text=os.path.basename(caminho))

def iniciar_processo():
    caminho_excel = caminho_var.get()
    if not caminho_excel:
        messagebox.showwarning("Aviso", "Selecione um arquivo Excel primeiro.")
        return

    btn_iniciar.config(state="disabled")
    log_box.delete(1.0, tk.END)
    progress_var.set(0)

    thread = threading.Thread(target=processar_arquivo, args=(caminho_excel, progress_var, log_box, btn_iniciar))
    thread.start()


root = tk.Tk()
root.title("Verificador de ads.txt")
root.geometry("600x400")
root.resizable(False, False)

caminho_var = tk.StringVar()
progress_var = tk.DoubleVar()

frame = tk.Frame(root, padx=20, pady=20)
frame.pack(fill="both", expand=True)

titulo = tk.Label(frame, text="🧾 Verificador de ads.txt", font=("Segoe UI", 16, "bold"))
titulo.pack(pady=10)

btn_arquivo = tk.Button(frame, text="Selecionar arquivo Excel", command=escolher_arquivo)
btn_arquivo.pack(pady=5)

label_arquivo = tk.Label(frame, text="Nenhum arquivo selecionado", fg="gray")
label_arquivo.pack(pady=5)

btn_iniciar = tk.Button(frame, text="Iniciar verificação", command=iniciar_processo, bg="#3b82f6", fg="white", width=20)
btn_iniciar.pack(pady=10)

progress = ttk.Progressbar(frame, variable=progress_var, maximum=100, length=400)
progress.pack(pady=5)

log_box = tk.Text(frame, height=10, width=60, wrap="word", font=("Consolas", 9))
log_box.pack(pady=10)

root.mainloop()
